#!/usr/bin/env python3
"""追加一行到 企业岗位汇总表.xlsx（文件不存在则新建）。

用法示例：
  python update_summary.py --company "某公司" --position "供应链运营" \
      --score 82 --plus "完善的应届生培训体系" --risk "加班偏高" \
      --advice "建议投递" --date "2026-08-21" [--path 企业岗位汇总表.xlsx]
"""
import argparse
import os

HEADERS = ["序号", "分析日期", "企业名称", "岗位名称",
           "综合评分(百分制)", "最大加分项", "最大风险项", "最终建议"]


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--company", required=True)
    p.add_argument("--position", required=True)
    p.add_argument("--score", required=True)
    p.add_argument("--plus", required=True)
    p.add_argument("--risk", required=True)
    p.add_argument("--advice", required=True)
    p.add_argument("--date", default="")
    p.add_argument("--path", default="企业岗位汇总表.xlsx")
    a = p.parse_args()

    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        raise SystemExit("缺少依赖 openpyxl，请先执行：pip install openpyxl")

    if os.path.exists(a.path):
        wb = load_workbook(a.path)
        ws = wb.active
        # 若首行不是标准表头，则补充表头行
        first = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        if first != HEADERS:
            ws.insert_rows(1)
            for col, h in enumerate(HEADERS, start=1):
                ws.cell(row=1, column=col, value=h)
        max_seq = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and isinstance(row[0], int):
                max_seq = max(max_seq, row[0])
        next_seq = max_seq + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        next_seq = 1

    ws.append([next_seq, a.date, a.company, a.position,
               a.score, a.plus, a.risk, a.advice])
    wb.save(a.path)
    print(f"OK: 已追加第 {next_seq} 行 -> {a.path}")


if __name__ == "__main__":
    main()
