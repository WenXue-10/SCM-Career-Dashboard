#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
企业岗位汇总表维护助手（含 ins 风精致样式）。

用法:
  # 追加一行（自动带样式）
  python append_summary.py <xlsx路径> \
      --company "比亚迪 BYD" \
      --role "计划专员（产能规划、供需平衡-电控）" \
      --score 68 \
      --pro "弗迪动力电控赛道+济南属地，成长与地域双优" \
      --risk "加班/轮岗未锁定，需面试承诺" \
      --advice "建议投递，面试必锁加班与轮岗承诺" \
      [--date 2026-08-21]

  # 更新已存在的公司记录（按企业名称匹配，不重复追加）
  python append_summary.py <xlsx路径> --update --company "比亚迪 BYD" \
      --role "..." --score 68 --pro "..." --risk "..." --advice "..."

  # 仅对已有 xlsx 重新应用精致样式（不改动数据）
  python append_summary.py <xlsx路径> --restyle

样式规范（ins 风，与 Word 报告一致）:
  - 表头: 干玫瑰粉填充 #B5838D + 白字加粗 12pt + 居中
  - 数据行: 斑马纹(白 / 极浅粉 #F7F0EE) + 微软雅黑 10.5pt + 自动换行 + 垂直居中
  - 边框: 浅陶土细线 #E8D5CD
  - 列宽: 序号6 / 日期13 / 企业16 / 岗位28 / 评分12 / 加分30 / 风险34 / 建议46(最宽)
  - 冻结首行
"""
import argparse
import datetime
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write("需要 openpyxl：请先 pip install openpyxl\n")
    raise SystemExit(2)

HEADERS = ["序号", "分析日期", "企业名称", "岗位名称",
           "综合评分(百分制)", "最大加分项", "最大风险项", "最终建议"]

# ---- ins 风配色 ----
HDR_FILL = "B5838D"      # 干玫瑰粉（表头）
HDR_TXT = "FFFFFF"       # 白字
ZEBRA = "F7F0EE"         # 极浅粉（偶数行）
BORDER = "E8D5CD"        # 浅陶土边框
FONT_NAME = "微软雅黑"

# 列宽（字符宽，建议列最宽）
COL_WIDTHS = {
    1: 6,    # 序号
    2: 13,   # 分析日期
    3: 16,   # 企业名称
    4: 28,   # 岗位名称
    5: 12,   # 综合评分
    6: 30,   # 最大加分项
    7: 34,   # 最大风险项
    8: 46,   # 最终建议（最宽）
}
CENTER_COLS = {1, 2, 5}  # 序号/日期/评分居中


def apply_summary_style(ws):
    """整表精致化：表头、斑马纹、边框、列宽、冻结、对齐（不改动数据）。"""
    ncol = len(HEADERS)
    # 确保表头存在
    if ws.cell(1, 1).value != HEADERS[0]:
        ws.insert_rows(1)
    for c in range(1, ncol + 1):
        ws.cell(1, c).value = HEADERS[c - 1]

    # 表头样式
    hdr_font = Font(name=FONT_NAME, size=12, bold=True, color=HDR_TXT)
    hdr_fill = PatternFill("solid", fgColor=HDR_FILL)
    thin = Side(style="thin", color=BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 数据行样式
    data_font = Font(name=FONT_NAME, size=10.5, color="333333")
    for r in range(2, ws.max_row + 1):
        zebra_fill = PatternFill("solid", fgColor=ZEBRA) if r % 2 == 0 else None
        for c in range(1, ncol + 1):
            cell = ws.cell(r, c)
            cell.font = data_font
            cell.border = border
            cell.fill = zebra_fill or PatternFill(fill_type=None)
            horiz = "center" if c in CENTER_COLS else "left"
            cell.alignment = Alignment(horizontal=horiz, vertical="center", wrap_text=True)

    # 列宽
    for c, w in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # 冻结首行 + 自动筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{ws.max_row}"
    ws.sheet_view.showGridLines = False


def find_company_row(ws, company):
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 3).value or "").find(company) >= 0:
            return r
    return None


def main():
    ap = argparse.ArgumentParser(description="企业岗位汇总表维护助手（含 ins 风样式）")
    ap.add_argument("xlsx", help="企业岗位汇总表.xlsx 路径")
    ap.add_argument("--company", help="企业名称")
    ap.add_argument("--role", help="岗位名称")
    ap.add_argument("--score", help="综合评分(百分制)")
    ap.add_argument("--pro", dest="pro", help="最大加分项")
    ap.add_argument("--risk", help="最大风险项")
    ap.add_argument("--advice", help="最终建议")
    ap.add_argument("--date", default=None, help="分析日期 YYYY-MM-DD，缺省为今天")
    ap.add_argument("--update", action="store_true", help="更新已存在公司记录而非追加")
    ap.add_argument("--restyle", action="store_true", help="仅重新应用样式，不改动数据")
    args = ap.parse_args()

    date_str = args.date or datetime.date.today().strftime("%Y-%m-%d")
    if args.score is not None:
        try:
            args.score = int(args.score)
        except (ValueError, TypeError):
            pass

    # 仅重新样式化
    if args.restyle:
        if not os.path.exists(args.xlsx):
            sys.stderr.write("文件不存在: %s\n" % args.xlsx)
            raise SystemExit(2)
        wb = openpyxl.load_workbook(args.xlsx)
        apply_summary_style(wb.active)
        wb.save(args.xlsx)
        print(f"已重新应用精致样式 -> {args.xlsx}")
        return

    # 新建或读取
    if os.path.exists(args.xlsx):
        wb = openpyxl.load_workbook(args.xlsx)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet"

    # 找最大序号
    max_seq = 0
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        v = row[0]
        if isinstance(v, (int, float)):
            max_seq = max(max_seq, int(v))

    if args.update and args.company:
        target = find_company_row(ws, args.company)
        if target is not None:
            if args.role is not None: ws.cell(target, 4, args.role)
            if args.score is not None: ws.cell(target, 5, args.score)
            if args.pro is not None: ws.cell(target, 6, args.pro)
            if args.risk is not None: ws.cell(target, 7, args.risk)
            if args.advice is not None: ws.cell(target, 8, args.advice)
            if args.date is not None: ws.cell(target, 2, args.date)
            print(f"已更新第 {target} 行({args.company})")
        else:
            print("未找到匹配公司，改为追加")
            target = None
    else:
        target = None

    if target is None:
        next_seq = max_seq + 1
        if args.date is None:
            args.date = date_str
        ws.append([next_seq, args.date, args.company, args.role,
                   args.score, args.pro, args.risk, args.advice])
        print(f"已追加第 {next_seq} 行 -> {args.xlsx}")

    apply_summary_style(ws)
    wb.save(args.xlsx)
    print(f"样式已应用 -> {args.xlsx}")


if __name__ == "__main__":
    main()
